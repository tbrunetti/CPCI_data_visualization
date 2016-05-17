import sys
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.cell import get_column_letter, column_index_from_string
import os
from collections import Counter

#in order to run either function, clinical data must be available!

#first line of clincal data is header
def categorize(pathToClinicalData, pathToSIFTandCADDmatchedFile):
	#loads .xlsx clinical data
	wb=load_workbook(pathToClinicalData)
	#extracts "Sheet1" as working spreadsheet, can be changed to any sheet name listed in workbook
	workingSheet=wb.get_sheet_by_name("Sheet1")

	#will get index number for column 'AK' which is the column of diagnosis
	#assuming using spreadsheet with schedule column removed
	typeIndex=column_index_from_string('AK')
	#this is the study ID column in the spreadsheet, so it can match patientID with diagnosis
	patientIndex=column_index_from_string('A')

	#patientDiagnosis=patientID (key) with cancer classification(value)
	patientDiagnosis={}
	#value for can be adjusted to number of clinical records available
	for i in range(1, 234):
		#**NOTE**:dictionary key must be an int NOT string because object type is listed as a long from openpyxl
		patientDiagnosis[workingSheet.cell(row=i, column=patientIndex).value]=workingSheet.cell(row=i, column=typeIndex).value
	

	#only stores patients that have available genetic data key=patientID, value=diagnosis
	dataAvailable={}
	#stores patient dianosis only, for quick counting of patient dianosis
	dianosisInStudy=[]
	for files in os.listdir(pathToSIFTandCADDmatchedFile):
		studyID=''
		for i in range(0, len(files)):
			if files[i] =='-':
				break;
			else:
				studyID=studyID+str(files[i])
		
		dataAvailable[studyID]=patientDiagnosis[int(studyID)]
		dianosisInStudy.append(patientDiagnosis[int(studyID)])		

	print 'The total number of patients with data available is '+str(len(dataAvailable))
	print 'The number of patients with genetic data and dianosis are as follows: '+'\n'+str(Counter(dianosisInStudy))

	return dataAvailable

#input is output of formatting_data_for_pathology_dept.py
#or input can be a tab-delimited file as long as first column of each line is studyID/patientID
#output is two files, one organized by PDAC only patients, and the other is all other diagnosis 
def sortGeneticData(dataAvailable):
	f=open('PDAC-genetic-variants.txt', 'w')
	f2=open('non-PDAC-genetic-variants.txt', 'w')
		
	with open(sys.argv[1]) as input:
		headers=next(input)
		headers=headers.split('\t')
		for x in range(0, len(headers)):
			if x==(len(headers)-1):
				f.write(str(headers[x]))
				f2.write(str(headers[x]))
			else:
				f.write(str(headers[x])+'\t')
				f2.write(str(headers[x])+'\t')
		
		for line in input:
			line=line.split('\t')
			line[len(line)-1]=line[len(line)-1].rstrip('\n')
			if dataAvailable[line[0]]=='PDAC':
				for i in range(0, len(line)-1):
					f.write(str(line[i])+'\t')
				f.write(str(line[len(line)-1])+'\n')
			else:
				for i in range(0, len(line)-1):
					f2.write(str(line[i])+'\t')
				f2.write(str(line[len(line)-1])+'\n')


if __name__=='__main__':
	
	pathToClinicalData='/home/tonya/pan_can_project/panData_UCretro_DNA_mutations/Retrospective_panData_UC/RetroPancStudy_PathDatabase1_AS_newdeath_noPass_noScheduleCol.xlsx'
	#this path can be any list of files for which the first characters are the studyIDs for the patients, followed by a hyphen
	pathToSIFTandCADDmatchedFile='/home/tonya/pan_can_project/panData_UCretro_DNA_mutations/match-SIFT-CADD'
	
	#specifying categorize will only perform categorize function call
	#no specification run both function calls
	for x in range(len(sys.argv)):
		if sys.argv[1]=='categorize':
			dataAvailable=categorize(pathToClinicalData, pathToSIFTandCADDmatchedFile)
		else:
			dataAvailable=categorize(pathToClinicalData, pathToSIFTandCADDmatchedFile)
			sortGeneticData(dataAvailable)
	